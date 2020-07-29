#
# convert VHH excel-format vocabularies to hierarchical xml in collective-access format
#
import argparse
import os
import sys
import glob
from tqdm import tqdm
sys.path.append(os.getcwd())

from openpyxl import load_workbook
from collections import defaultdict

#
# config
#
parser = argparse.ArgumentParser()

parser.add_argument('--excel-file', action='store', dest='excel_files',
                    help='the vhh-vocab excel file location', required=True)

parser.add_argument('--out-file-hierarchy', action='store', dest='out_file',
                    help='xml output file destination', required=True)

args = parser.parse_args()


#
# parse excel
# 

excel_files = glob.glob(args.excel_files)
vocab_flat_global = {}
vocab_hierarchy_global = {}

for file in excel_files:

    wb = load_workbook(file)

    print("File:", file)
    #print("Found sheets:", wb.sheetnames)

    for sheet in wb:
        if sheet.sheet_state != 'visible':
            continue
        print(sheet.title)
        
        headings = list(sheet.iter_rows(min_row=1,max_row=1,max_col=10,values_only=True))[0]

        for row in sheet.iter_rows(min_row=2,values_only=True,max_col=10):

            if row[0] == None:
                continue

            id_ = sheet.title + "_" + str(row[0])

            vocab_flat_global[id_] = row

            hierarchy_loc = vocab_hierarchy_global
            for cat in row[1:5]:
                if cat not in hierarchy_loc:
                    hierarchy_loc[cat] = {}
                hierarchy_loc = hierarchy_loc[cat]

            if row[5] not in hierarchy_loc:
                hierarchy_loc[row[5]] = []
            hierarchy_loc = hierarchy_loc[row[5]]

            hierarchy_loc.append({"id":id_,headings[6]:row[6],headings[7]:row[7],headings[8]:row[8],headings[9]:row[9]})

#
# output
# 
rank=0

def make_xml(ind,tree_node,tree_node_data):
    global rank
    xml = ""

    skip = tree_node == None or tree_node == "x"

    ind_s = ' ' * ind
    ind_next = ind if skip else ind+4
    
    if not skip: rank+=1
    if not skip: xml+='\
'+ind_s+'<item idno="h_'+str(rank)+'" enabled="1" default="1" value="'+tree_node+'" rank="'+str(rank)+'" type="hierarchy name">\n\
'+ind_s+'    <labels>\n\
'+ind_s+'        <label locale="en_US" preferred="1">\n\
'+ind_s+'            <name_singular>'+tree_node+'</name_singular>\n\
'+ind_s+'            <name_plural>'+tree_node+'</name_plural>\n\
'+ind_s+'            <description/>\n\
'+ind_s+'        </label>\n\
'+ind_s+'    </labels>\n\
'+ind_s+'    <items>\n'

    if type(tree_node_data) == dict:

        for node,data in tree_node_data.items():
            xml += make_xml(ind_next,node,data)

    elif type(tree_node_data) == list:

        for data in tree_node_data:
            rank+=1

            xml+='\
'+ind_s+'       <item idno="'+data["id"]+'" enabled="1" default="1" value="'+data["Term_en"].lower().replace("\"","")+'" rank="'+str(rank)+'" type="concept">\n\
'+ind_s+'           <labels>\n\
'+ind_s+'               <label locale="en_US" preferred="1">\n\
'+ind_s+'                   <name_singular>'+data["Term_en"]+'</name_singular>\n\
'+ind_s+'                   <name_plural>'+data["Term_en"]+'</name_plural>\n\
'+ind_s+'                   <description>'+data["Term_de"]+" "+data["VHH_WikiData_LINK"]+'</description>\n\
'+ind_s+'               </label>\n\
'+ind_s+'           </labels>\n\
'+ind_s+'       </item>\n'

    if not skip: xml+='\
'+ind_s+'    </items>\n\
'+ind_s+'</item>\n'

    return xml

with open(args.out_file,"w",encoding="utf8") as out_file:

    xml = '<?xml version="1.0"?>\n<list code="vhh_vocabulary" hierarchical="1" system="0" vocabulary="1">\n    <items>\n'
    for top_node,data in vocab_hierarchy_global.items():

        xml += make_xml(4,top_node,data)

    xml += '    </items>\n</list>\n'

    out_file.write(xml)

