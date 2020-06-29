
import argparse
import os
import sys
import glob
from tqdm import tqdm
sys.path.append(os.getcwd())

from openpyxl import load_workbook
import json
from collections import defaultdict

#
# config
#
parser = argparse.ArgumentParser()

parser.add_argument('--excel-file', action='store', dest='excel_files',
                    help='the vhh-vocab excel file location', required=True)

parser.add_argument('--out-file-plain', action='store', dest='out_file_plain',
                    help='plain output file destination', required=True)

parser.add_argument('--out-file-flat', action='store', dest='out_file_flat',
                    help='flat, but full file destination', required=True)

parser.add_argument('--out-file-hierarchy', action='store', dest='out_file',
                    help='json output file destination', required=True)

args = parser.parse_args()


#
# parse excel
# 

excel_files = glob.glob(args.excel_files)
vocab_flat_global = {}
vocab_hierarchy_global = {}

for file in excel_files:

    wb = load_workbook(file)

    print("File:", file)
    #print("Found sheets:", wb.sheetnames)

    for sheet in wb:
        if sheet.sheet_state != 'visible':
            continue
        print(sheet.title)
        
        headings = list(sheet.iter_rows(min_row=1,max_row=1,max_col=10,values_only=True))[0]

        for row in sheet.iter_rows(min_row=2,values_only=True,max_col=10):

            if row[0] == None:
                continue

            id_ = sheet.title + "_" + str(row[0])

            vocab_flat_global[id_] = row

            hierarchy_loc = vocab_hierarchy_global
            for cat in row[1:5]:
                if cat not in hierarchy_loc:
                    hierarchy_loc[cat] = {}
                hierarchy_loc = hierarchy_loc[cat]

            if row[5] not in hierarchy_loc:
                hierarchy_loc[row[5]] = []
            hierarchy_loc = hierarchy_loc[row[5]]

            hierarchy_loc.append({"id":id_,headings[6]:row[6],headings[7]:row[7],headings[8]:row[8],headings[9]:row[9]})

#
# output
# 

with open(args.out_file_flat,"w",encoding="utf8") as out_file_flat, \
     open(args.out_file_plain,"w",encoding="utf8") as out_file_plain:

    for id_,row in vocab_flat_global.items():
        out_file_flat.write(id_+"\t"+"\t".join([str(s) if s!=None else "" for s in row])+"\n")
        out_file_plain.write(id_+"\t"+(row[7] if row[7] != None else "")+"\n")

with open(args.out_file,"w",encoding="utf8") as out_file:
    json.dump(vocab_hierarchy_global,out_file,indent=2)
